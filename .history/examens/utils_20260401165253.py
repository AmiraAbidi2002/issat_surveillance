import openpyxl
from datetime import datetime, time, date
from users.models import User
from .models import Examen


# =========================================================
# OUTILS COMMUNS
# =========================================================
def parse_time(value):
    """Convertit différents formats d'heure en objet time."""
    if isinstance(value, time):
        return value

    if isinstance(value, datetime):
        return value.time()

    if isinstance(value, str):
        for fmt in ["%H:%M:%S", "%H:%M"]:
            try:
                return datetime.strptime(value.strip(), fmt).time()
            except ValueError:
                continue

    raise ValueError(f"Format d'heure invalide : {value}")


def parse_date(value):
    """Convertit différents formats de date en objet date."""
    if isinstance(value, date) and not isinstance(value, datetime):
        return value

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, str):
        for fmt in ["%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"]:
            try:
                return datetime.strptime(value.strip(), fmt).date()
            except ValueError:
                continue

    raise ValueError(f"Format de date invalide : {value}")


def find_enseignant(nom_enseignant):
    """
    Recherche un enseignant par :
    - username exact
    - last_name
    - username reconstruit (Ben Ali -> ben_ali)
    """
    nom_enseignant = str(nom_enseignant).strip()

    # 1) username exact
    user = User.objects.filter(
        username__iexact=nom_enseignant,
        role="enseignant"
    ).first()

    # 2) nom de famille
    if not user:
        user = User.objects.filter(
            last_name__iexact=nom_enseignant,
            role="enseignant"
        ).first()

    # 3) nom transformé en username
    if not user:
        username_guess = nom_enseignant.lower().replace(" ", "_")
        user = User.objects.filter(
            username__iexact=username_guess,
            role="enseignant"
        ).first()

    if not user:
        return None

    return user.profil_enseignant


# =========================================================
# IMPORT EXAMENS
# =========================================================
def importer_examens(fichier):
    wb = openpyxl.load_workbook(fichier, data_only=True)

    if "Examens" not in wb.sheetnames:
        return {
            "success": False,
            "erreur": "La feuille 'Examens' est introuvable"
        }

    ws = wb["Examens"]

    headers = [
        str(cell.value).strip().lower() if cell.value else ""
        for cell in ws[1]
    ]

    mapping = {
        "date": "date_exam",
        "responsable": "enseignant_username"
    }
    headers = [mapping.get(h, h) for h in headers]

    colonnes_requises = {
        "matiere",
        "classe",
        "date_exam",
        "heure_debut",
        "heure_fin",
        "salle",
        "nb_etudiants",
        "enseignant_username",
    }

    manquantes = colonnes_requises - set(headers)
    if manquantes:
        return {
            "success": False,
            "erreur": f"Colonnes manquantes : {', '.join(manquantes)}"
        }

    crees = []
    erreurs = []

    for num_ligne, row in enumerate(
        ws.iter_rows(min_row=2, values_only=True),
        start=2
    ):
        if all(v is None for v in row):
            continue

        data = dict(zip(headers, row))

        try:
            nom_enseignant = data.get("enseignant_username")
            enseignant = None

            if nom_enseignant:
                enseignant = find_enseignant(nom_enseignant)

                if not enseignant:
                    erreurs.append(
                        f"Ligne {num_ligne} : enseignant '{nom_enseignant}' introuvable"
                    )
                    continue

            examen, created = Examen.objects.update_or_create(
                matiere=str(data["matiere"]).strip(),
                classe=str(data["classe"]).strip(),
                date_exam=parse_date(data["date_exam"]),
                defaults={
                    "heure_debut": parse_time(data["heure_debut"]),
                    "heure_fin": parse_time(data["heure_fin"]),
                    "salle": str(data["salle"]).strip(),
                    "nb_etudiants": int(data["nb_etudiants"]),
                    "enseignant_responsable": enseignant,
                },
            )

            crees.append({
                "id": examen.id,
                "matiere": examen.matiere,
                "classe": examen.classe,
                "date": str(examen.date_exam),
                "action": "créé" if created else "mis à jour",
            })

        except Exception as e:
            erreurs.append(f"Ligne {num_ligne} : {str(e)}")

    return {
        "success": True,
        "total": len(crees),
        "crees": crees,
        "erreurs": erreurs,
    }


# =========================================================
# IMPORT DISPONIBILITES
# =========================================================
def importer_disponibilites(fichier):
    from disponibilites.models import Disponibilite

    wb = openpyxl.load_workbook(fichier, data_only=True)

    if "Disponibilites" not in wb.sheetnames:
        return {
            "success": False,
            "erreur": "La feuille 'Disponibilites' est introuvable"
        }

    ws = wb["Disponibilites"]

    headers = [
        str(cell.value).strip().lower() if cell.value else ""
        for cell in ws[1]
    ]

    mapping = {
        "date": "date_dispo",
        "nom_enseignant": "enseignant_username"
    }
    headers = [mapping.get(h, h) for h in headers]

    crees = []
    erreurs = []

    for num_ligne, row in enumerate(
        ws.iter_rows(min_row=2, values_only=True),
        start=2
    ):
        if all(v is None for v in row):
            continue

        data = dict(zip(headers, row))

        try:
            enseignant = find_enseignant(data["enseignant_username"])

            if not enseignant:
                erreurs.append(
                    f"Ligne {num_ligne} : enseignant '{data['enseignant_username']}' introuvable"
                )
                continue

            dispo, created = Disponibilite.objects.update_or_create(
                enseignant=enseignant,
                date_dispo=parse_date(data["date_dispo"]),
                heure_debut=parse_time(data["heure_debut"]),
                defaults={
                    "heure_fin": parse_time(data["heure_fin"])
                }
            )

            crees.append({
                "enseignant": data["enseignant_username"],
                "date": str(dispo.date_dispo),
                "action": "créé" if created else "mis à jour",
            })

        except Exception as e:
            erreurs.append(f"Ligne {num_ligne} : {str(e)}")

    return {
        "success": True,
        "total": len(crees),
        "crees": crees,
        "erreurs": erreurs,
    }


# =========================================================
# IMPORT ABSENCES
# =========================================================
def importer_absenteisme(fichier):
    wb = openpyxl.load_workbook(fichier, data_only=True)

    if "Absences" not in wb.sheetnames:
        return {
            "success": False,
            "erreur": "La feuille 'Absences' est introuvable"
        }

    ws = wb["Absences"]

    headers = [
        str(cell.value).strip().lower() if cell.value else ""
        for cell in ws[1]
    ]

    mapping = {
        "nom_enseignant": "enseignant_username",
        "heures_absence": "heures_manquees"
    }
    headers = [mapping.get(h, h) for h in headers]

    traites = []
    erreurs = []

    for num_ligne, row in enumerate(
        ws.iter_rows(min_row=2, values_only=True),
        start=2
    ):
        if all(v is None for v in row):
            continue

        data = dict(zip(headers, row))

        try:
            enseignant = find_enseignant(data["enseignant_username"])

            if not enseignant:
                erreurs.append(
                    f"Ligne {num_ligne} : enseignant '{data['enseignant_username']}' introuvable"
                )
                continue

            heures = float(data["heures_manquees"])
            enseignant.heures_surveillance_dues += heures
            enseignant.save()

            traites.append({
                "enseignant": data["enseignant_username"],
                "heures_dues": enseignant.heures_surveillance_dues,
                "action": "mis à jour",
            })

        except Exception as e:
            erreurs.append(f"Ligne {num_ligne} : {str(e)}")

    return {
        "success": True,
        "total": len(traites),
        "traites": traites,
        "erreurs": erreurs,
    }