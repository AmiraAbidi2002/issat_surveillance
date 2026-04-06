import openpyxl
from datetime import datetime, time, date
from users.models import User
from .models import Examen


# =========================================================
# OUTILS COMMUNS
# =========================================================
def parse_time(value):
    """Convertit différents formats d'heure en objet time."""
    if isinstance(value, time):
        return value

    if isinstance(value, datetime):
        return value.time()

    if isinstance(value, str):
        for fmt in ["%H:%M:%S", "%H:%M"]:
            try:
                return datetime.strptime(value.strip(), fmt).time()
            except ValueError:
                continue

    raise ValueError(f"Format d'heure invalide : {value}")


def parse_date(value):
    """Convertit différents formats de date en objet date."""
    if isinstance(value, date) and not isinstance(value, datetime):
        return value

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, str):
        for fmt in ["%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"]:
            try:
                return datetime.strptime(value.strip(), fmt).date()
            except ValueError:
                continue

    raise ValueError(f"Format de date invalide : {value}")


def find_enseignant(nom_enseignant):
    """
    Recherche un enseignant par :
    - username exact
    - last_name
    - username reconstruit (Ben Ali -> ben_ali)
    """
    nom_enseignant = str(nom_enseignant).strip()

    # 1) username exact
    user = User.objects.filter(
        username__iexact=nom_enseignant,
        role="enseignant"
    ).first()

    # 2) nom de famille
    if not user:
        user = User.objects.filter(
            last_name__iexact=nom_enseignant,
            role="enseignant"
        ).first()

    # 3) nom transformé en username
    if not user:
        username_guess = nom_enseignant.lower().replace(" ", "_")
        user = User.objects.filter(
            username__iexact=username_guess,
            role="enseignant"
        ).first()

    if not user:
        return None

    return user.profil_enseignant


# =========================================================
# IMPORT EXAMENS
# =========================================================
def importer_examens(fichier):
    wb = openpyxl.load_workbook(fichier, data_only=True)

    if "Examens" not in wb.sheetnames:
        return {
            "success": False,
            "erreur": "La feuille 'Examens' est introuvable"
        }

    ws = wb["Examens"]

    headers = [
        str(cell.value).strip().lower() if cell.value else ""
        for cell in ws[1]
    ]

    mapping = {
        "date": "date_exam",
        "responsable": "enseignant_username"
    }
    headers = [mapping.get(h, h) for h in headers]

    colonnes_requises = {
        "matiere",
        "classe",
        "date_exam",
        "heure_debut",
        "heure_fin",
        "salle",
        "nb_etudiants",
        "enseignant_username",
    }

    manquantes = colonnes_requises - set(headers)
    if manquantes:
        return {
            "success": False,
            "erreur": f"Colonnes manquantes : {', '.join(manquantes)}"
        }

    crees = []
    erreurs = []

    for num_ligne, row in enumerate(
        ws.iter_rows(min_row=2, values_only=True),
        start=2
    ):
        if all(v is None for v in row):
            continue

        data = dict(zip(headers, row))

        try:
            nom_enseignant = data.get("enseignant_username")
            enseignant = None

            if nom_enseignant:
                enseignant = find_enseignant(nom_enseignant)

                if not enseignant:
                    erreurs.append(
                        f"Ligne {num_ligne} : enseignant '{nom_enseignant}' introuvable"
                    )
                    continue

            examen, created = Examen.objects.update_or_create(
                matiere=str(data["matiere"]).strip(),
                classe=str(data["classe"]).strip(),
                date_exam=parse_date(data["date_exam"]),
                defaults={
                    "heure_debut": parse_time(data["heure_debut"]),
                    "heure_fin": parse_time(data["heure_fin"]),
                    "salle": str(data["salle"]).strip(),
                    "nb_etudiants": int(data["nb_etudiants"]),
                    "enseignant_responsable": enseignant,
                },
            )

            crees.append({
                "id": examen.id,
                "matiere": examen.matiere,
                "classe": examen.classe,
                "date": str(examen.date_exam),
                "action": "créé" if created else "mis à jour",
            })

        except Exception as e:
            erreurs.append(f"Ligne {num_ligne} : {str(e)}")

    return {
        "success": True,
        "total": len(crees),
        "crees": crees,
        "erreurs": erreurs,
    }


# =========================================================
# IMPORT DISPONIBILITES
# =========================================================
def importer_disponibilites(fichier):
    from disponibilites.models import Disponibilite

    wb = openpyxl.load_workbook(fichier, data_only=True)

    if "Disponibilites" not in wb.sheetnames:
        return {
            "success": False,
            "erreur": "La feuille 'Disponibilites' est introuvable"
        }

    ws = wb["Disponibilites"]

    headers = [
        str(cell.value).strip().lower() if cell.value else ""
        for cell in ws[1]
    ]

    mapping = {
        "date": "date_dispo",
        "nom_enseignant": "enseignant_username"
    }
    headers = [mapping.get(h, h) for h in headers]

    crees = []
    erreurs = []

    for num_ligne, row in enumerate(
        ws.iter_rows(min_row=2, values_only=True),
        start=2
    ):
        if all(v is None for v in row):
            continue

        data = dict(zip(headers, row))

        try:
            enseignant = find_enseignant(data["enseignant_username"])

            if not enseignant:
                erreurs.append(
                    f"Ligne {num_ligne} : enseignant '{data['enseignant_username']}' introuvable"
                )
                continue

            dispo, created = Disponibilite.objects.update_or_create(
                enseignant=enseignant,
                date_dispo=parse_date(data["date_dispo"]),
                heure_debut=parse_time(data["heure_debut"]),
                defaults={
                    "heure_fin": parse_time(data["heure_fin"])
                }
            )

            crees.append({
                "enseignant": data["enseignant_username"],
                "date": str(dispo.date_dispo),
                "action": "créé" if created else "mis à jour",
            })

        except Exception as e:
            erreurs.append(f"Ligne {num_ligne} : {str(e)}")

    return {
        "success": True,
        "total": len(crees),
        "crees": crees,
        "erreurs": erreurs,
    }


# =========================================================
# IMPORT ABSENCES
# =========================================================
def importer_absenteisme(fichier):
    wb = openpyxl.load_workbook(fichier, data_only=True)

    if "Absences" not in wb.sheetnames:
        return {
            "success": False,
            "erreur": "La feuille 'Absences' est introuvable"
        }

    ws = wb["Absences"]

    headers = [
        str(cell.value).strip().lower() if cell.value else ""
        for cell in ws[1]
    ]

    mapping = {
        "nom_enseignant": "enseignant_username"
    }
    headers = [mapping.get(h, h) for h in headers]

    colonnes_requises = {
        "enseignant_username",
        "heure_debut",
        "heure_fin",
    }

    manquantes = colonnes_requises - set(headers)
    if manquantes:
        return {
            "success": False,
            "erreur": f"Colonnes manquantes : {', '.join(manquantes)}"
        }

    traites = []
    erreurs = []

    for num_ligne, row in enumerate(
        ws.iter_rows(min_row=2, values_only=True),
        start=2
    ):
        if all(v is None for v in row):
            continue

        data = dict(zip(headers, row))

        try:
            enseignant = find_enseignant(data["enseignant_username"])

            if not enseignant:
                erreurs.append(
                    f"Ligne {num_ligne} : enseignant '{data['enseignant_username']}' introuvable"
                )
                continue

            heure_debut = parse_time(data["heure_debut"])
            heure_fin = parse_time(data["heure_fin"])

            # calcul durée absence
            dt_debut = datetime.combine(date.today(), heure_debut)
            dt_fin = datetime.combine(date.today(), heure_fin)

            heures = (dt_fin - dt_debut).seconds / 3600

            enseignant.heures_surveillance_dues += heures
            enseignant.save()

            traites.append({
                "enseignant": data["enseignant_username"],
                "heures_ajoutees": heures,
                "heures_dues": enseignant.heures_surveillance_dues,
                "action": "mis à jour",
            })

        except Exception as e:
            erreurs.append(f"Ligne {num_ligne} : {str(e)}")

    return {
        "success": True,
        "total": len(traites),
        "traites": traites,
        "erreurs": erreurs,
    }

def importer_fichier_complet(fichier):
    """
    Import unique du fichier Excel complet.
    Traite les 4 feuilles dans l'ordre :
      1. Enseignants  → crée User + Enseignant + Departement
      2. Examens      → crée Examen
      3. Disponibilites → crée Disponibilite
      4. Absences     → met à jour heures_surveillance_dues
    Retourne un rapport détaillé.
    """
    import openpyxl
    from users.models import User, Departement, Enseignant
    from disponibilites.models import Disponibilite
    from .models import Examen

    wb = openpyxl.load_workbook(fichier, data_only=True)
    feuilles = [s.lower() for s in wb.sheetnames]

    rapport = {
        'departements': {'crees': 0, 'existants': 0},
        'enseignants':  {'crees': 0, 'existants': 0, 'erreurs': []},
        'examens':      {'crees': 0, 'maj': 0,       'erreurs': []},
        'disponibilites':{'crees': 0, 'maj': 0,      'erreurs': []},
        'absences':     {'traites': 0,                'erreurs': []},
    }

    # ══════════════════════════════════════════════════════
    # FEUILLE 1 — Enseignants
    # ══════════════════════════════════════════════════════
    ws_ens = _get_sheet(wb, 'enseignants')
    if ws_ens is None:
        return {'success': False,
                'erreur': "Feuille 'Enseignants' introuvable."}

    headers = _headers(ws_ens)
    required = {'nom_complet', 'username', 'departement',
                'heures_semaine', 'mot_de_passe'}
    manquantes = required - set(headers)
    if manquantes:
        return {'success': False,
                'erreur': f"Colonnes manquantes dans Enseignants : {manquantes}"}

    for num, row in enumerate(ws_ens.iter_rows(min_row=2, values_only=True), 2):
        if all(v is None for v in row):
            continue
        data = dict(zip(headers, row))
        nom_complet = str(data.get('nom_complet', '') or '').strip()
        username    = str(data.get('username',    '') or '').strip()
        if not username:
            continue

        try:
            # Département
            nom_dept = str(data.get('departement', 'Informatique') or 'Informatique').strip()
            dept, created = Departement.objects.get_or_create(nom=nom_dept)
            if created:
                rapport['departements']['crees'] += 1
            else:
                rapport['departements']['existants'] += 1

            # Extraire prénom / nom de famille
            first_name, last_name = _split_nom(nom_complet)

            heures = float(data.get('heures_semaine') or 9)
            email  = str(data.get('email') or f"{username}@issat.rnu.tn").strip()
            mdp    = str(data.get('mot_de_passe') or 'Issat2025!').strip()

            # User
            user, u_created = User.objects.get_or_create(
                username=username,
                defaults={
                    'first_name': first_name,
                    'last_name':  last_name,
                    'email':      email,
                    'role':       'enseignant',
                }
            )
            if u_created:
                user.set_password(mdp)
                user.save()

            # Enseignant
            _, e_created = Enseignant.objects.get_or_create(
                user=user,
                defaults={
                    'departement':              dept,
                    'heures_enseignement':      heures,
                    'heures_surveillance_dues': heures,
                    'heures_effectuees':        0.0,
                }
            )

            if u_created and e_created:
                rapport['enseignants']['crees'] += 1
            else:
                rapport['enseignants']['existants'] += 1

        except Exception as e:
            rapport['enseignants']['erreurs'].append(
                f"Ligne {num} ({username}) : {str(e)}"
            )

    # ══════════════════════════════════════════════════════
    # FEUILLE 2 — Examens
    # ══════════════════════════════════════════════════════
    ws_ex = _get_sheet(wb, 'examens')
    if ws_ex:
        headers = _headers(ws_ex)
        for num, row in enumerate(ws_ex.iter_rows(min_row=2, values_only=True), 2):
            if all(v is None for v in row):
                continue
            data = dict(zip(headers, row))
            try:
                username_resp = str(data.get('enseignant_username') or '').strip()
                enseignant_resp = None
                if username_resp:
                    try:
                        u = User.objects.get(username=username_resp)
                        enseignant_resp = u.profil_enseignant
                    except (User.DoesNotExist, Enseignant.DoesNotExist):
                        rapport['examens']['erreurs'].append(
                            f"Ligne {num} : enseignant '{username_resp}' introuvable"
                        )

                examen, created = Examen.objects.update_or_create(
                    matiere   = str(data['matiere']).strip(),
                    classe    = str(data['classe']).strip(),
                    date_exam = parse_date(data['date_exam']),
                    defaults={
                        'heure_debut':            parse_time(data['heure_debut']),
                        'heure_fin':              parse_time(data['heure_fin']),
                        'salle':                  str(data.get('salle', 'TBD')).strip(),
                        'nb_etudiants':           int(data.get('nb_etudiants') or 30),
                        'enseignant_responsable': enseignant_resp,
                    }
                )
                if created:
                    rapport['examens']['crees'] += 1
                else:
                    rapport['examens']['maj'] += 1

            except Exception as e:
                rapport['examens']['erreurs'].append(f"Ligne {num} : {str(e)}")

    # ══════════════════════════════════════════════════════
    # FEUILLE 3 — Disponibilites
    # ══════════════════════════════════════════════════════
    ws_dis = _get_sheet(wb, 'disponibilites')
    if ws_dis:
        headers = _headers(ws_dis)
        for num, row in enumerate(ws_dis.iter_rows(min_row=2, values_only=True), 2):
            if all(v is None for v in row):
                continue
            data = dict(zip(headers, row))
            try:
                username = str(data.get('enseignant_username') or '').strip()
                if not username:
                    continue
                user = User.objects.get(username=username)
                ens  = user.profil_enseignant

                dispo, created = Disponibilite.objects.update_or_create(
                    enseignant  = ens,
                    date_dispo  = parse_date(data['date_dispo']),
                    heure_debut = parse_time(data['heure_debut']),
                    defaults={'heure_fin': parse_time(data['heure_fin'])}
                )
                if created:
                    rapport['disponibilites']['crees'] += 1
                else:
                    rapport['disponibilites']['maj'] += 1

            except User.DoesNotExist:
                rapport['disponibilites']['erreurs'].append(
                    f"Ligne {num} : enseignant '{username}' introuvable"
                )
            except Exception as e:
                rapport['disponibilites']['erreurs'].append(
                    f"Ligne {num} : {str(e)}"
                )

    # ══════════════════════════════════════════════════════
    # FEUILLE 4 — Absences
    # ══════════════════════════════════════════════════════
    ws_abs = _get_sheet(wb, 'absences')
    if ws_abs:
        headers = _headers(ws_abs)
        for num, row in enumerate(ws_abs.iter_rows(min_row=2, values_only=True), 2):
            if all(v is None for v in row):
                continue
            data = dict(zip(headers, row))
            try:
                username = str(data.get('enseignant_username') or '').strip()
                if not username:
                    continue
                user = User.objects.get(username=username)
                ens  = user.profil_enseignant

                # Calculer la durée de l'absence
                from datetime import datetime
                h_debut = parse_time(data['heure_debut'])
                h_fin   = parse_time(data['heure_fin'])
                d1 = datetime.combine(datetime.today(), h_debut)
                d2 = datetime.combine(datetime.today(), h_fin)
                heures_manquees = max(0.0, (d2 - d1).seconds / 3600)

                # Ajouter les heures manquées aux heures dues
                ens.heures_surveillance_dues = (
                    ens.heures_enseignement + heures_manquees
                )
                ens.save()
                rapport['absences']['traites'] += 1

            except User.DoesNotExist:
                rapport['absences']['erreurs'].append(
                    f"Ligne {num} : enseignant '{username}' introuvable"
                )
            except Exception as e:
                rapport['absences']['erreurs'].append(
                    f"Ligne {num} : {str(e)}"
                )

    rapport['success'] = True
    return rapport


# ── Helpers privés ────────────────────────────────────────────────

def _get_sheet(wb, nom_recherche):
    """Trouve une feuille par nom (insensible à la casse)."""
    for name in wb.sheetnames:
        if name.lower() == nom_recherche.lower():
            return wb[name]
    return None


def _headers(ws):
    """Retourne les en-têtes de la ligne 1 en minuscules."""
    return [
        str(cell.value).strip().lower().replace(' ', '_')
        if cell.value else ''
        for cell in next(ws.iter_rows(max_row=1))
    ]


def _split_nom(nom_complet):
    """
    Extrait prénom et nom depuis 'NOM_FAMILLE Prenom'.
    Convention : les mots tout en majuscules = nom de famille.
    """
    parts = str(nom_complet).strip().split()
    noms_famille = [
        p for p in parts
        if p.isupper() or (len(p) > 2 and p[1:].isupper())
    ]
    prenoms = [p for p in parts if p not in noms_famille]
    first_name = ' '.join(prenoms)  if prenoms      else (parts[-1] if parts else '')
    last_name  = ' '.join(noms_famille) if noms_famille else (parts[0]  if parts else '')
    return first_name, last_name